from regex import P
import xlsxwriter
from openpyxl import load_workbook
import ast

def name():
    wb = load_workbook('WS_FINAL3.xlsx')
    ws = wb.active
    name = []
    FirstName = []
    LastName = []
    Lwynb = []
    for i in range(1,1332):
        name.append(ws.cell(row=i,column=1).value)
    for i in name:
        namestr = i.split(' ')
        FirstName.append(namestr[0])
        Lwynb.append(namestr[-1])
        if namestr[-2] == 'Jr.' or namestr[-2] == 'Jr' or namestr[-2] == 'I' or namestr[-2] == 'II' or namestr[-2] == 'III' or namestr[-2] == 'IV' or namestr[-2] == 'V' or namestr[-2] == 'Sr.':
            LastName.append(namestr[-3]+' '+namestr[-2])
        else:
            LastName.append(namestr[-2])
    for idx,name in enumerate(LastName):
        for letter in name:
            if letter == ')':
                name = name.replace(')','')
                LastName[idx] = name
    wb.close()
    # print(LastName)
    return FirstName, LastName, Lwynb
# print(name())

def splitaddress():
    wb = load_workbook('WS_FINAL3.xlsx')
    ws = wb.active
    address = []
    Addressline = []
    city = []
    state = []
    zipcode = []
    for i in range(1,1332):
        address.append(ws.cell(row=i,column=3).value)
    for i in address:
        addressstr = i.split(',')
        # print(addressstr)
        Addressline.append(''.join(addressstr[0:-2]))
        city.append(addressstr[-2])
        statestr = addressstr[-1].split(' ')
        state.append(statestr[1])
        zipcode.append(statestr[2])
    # print(Addressline)
    # print(city)
    # print(state)
    # print(zipcode)
    wb.close()
    return Addressline, city, state, zipcode
print(splitaddress()[3])

def splitphone():
    wb = load_workbook('WS_FINAL3.xlsx')
    ws = wb.active
    line = []
    phone = []
    fax = []
    for i in range(1,1332):
        line.append(ws.cell(row=i,column=4).value)
    # print(line)
    for i in line:
        # print(i)
        linestr = i.split('  |  ')
        # print(linestr)
        phone.append(linestr[0])
        fax.append(linestr[1])
    print(phone)
    print(fax)
    wb.close()
    return phone,fax

def splitemailweb():
    wb = load_workbook('WS_FINAL3.xlsx')
    ws = wb.active
    line = []
    email = []
    web = []
    for i in range(1,1332):
        line.append(ws.cell(row=i,column=5).value)
    for i in line:
        # print(i)
        linestr = i.split('  |  ')
        email.append(linestr[0])
        web.append(linestr[1])
    print(email)
    print(web)
    wb.close()
    return email,web


def append():
    dcwb= xlsxwriter.Workbook('WS_FINAL2.xlsx')
    dcws = dcwb.add_worksheet('infosheet')
    row = 0
    for fn in name()[0]:
        dcws.write(row,0,fn)
        row+=1
    row = 0
    for ls in name()[1]:
        dcws.write(row,1,ls)
        row+=1
    row = 0
    for nb in name()[2]:
        dcws.write(row,2,nb)
        row+=1
    row = 0
    for phone in splitphone()[0]:
        dcws.write(row,7,phone)
        row+=1
    row = 0
    for fax in splitphone()[1]:
        dcws.write(row,8,fax)
        row+=1
    row = 0
    for email in splitemailweb()[0]:
        dcws.write(row,9,email)
        row+=1
    row = 0
    for web in splitemailweb()[1]:
        dcws.write(row,10,web)
        row+=1
    row = 0
    for ad in splitaddress()[0]:
        dcws.write(row,3,ad)
        row+=1
    row = 0
    for ct in splitaddress()[1]:
        dcws.write(row,4,ct)
        row+=1
    row = 0
    for st in splitaddress()[2]:
        dcws.write(row,5,st)
        row+=1
    row = 0
    for zp in splitaddress()[3]:
        dcws.write(row,6,zp)
        row+=1
    dcwb.close()
append()

