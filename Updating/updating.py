import csv
from csv import writer
from openpyxl import load_workbook

def deleteables(docname):
    dset = set()
    file = open(f'{docname}', encoding='utf8')
    csvreader = csv.reader(file)
    a = 0
    for row in csvreader:
        dset.add(row[-2])
        a += 1
    print(f'total {a} in deleteables from {docname}')
    file.close()
    # print(dset)
    return dset


def alldoc(docname):
    file = open(f'{docname}', encoding='utf8')
    csvreader = csv.reader(file)
    rows = set() #all rows

    num_of_lines_in_all = 0

    #count lines in all.csv
    for all in csvreader:
        # print(all)
        rows.add(all[-2])
        num_of_lines_in_all += 1
    print(f'total {num_of_lines_in_all} in the alldoc')
    file.close()
    return rows

dset = deleteables('Updating/immigration_marketing_04152022.csv')
rows = alldoc('Updating/immigration_marketing_06192022.csv')
new = rows - dset
# print(dset)
# print(rows)
print(len(new))

def newdoc():
    file = open(f'Updating/immigration_marketing_06192022.csv', encoding='utf8')
    csvreader = csv.reader(file)
    nrows = [] #all rows

    num_of_lines_in_all = 0

    #count lines in all.csv
    for i in csvreader:
        # print(i[-2])
        for j in new: 
            if i[-2] == j:
                if i[-2] != '':
                        nrows.append(i)
                        num_of_lines_in_all += 1
    print(f'total {num_of_lines_in_all} in the newdoc')
    file.close()
    return nrows

nrows = newdoc()
with open('Updating/newdoc.csv', 'w') as f_object:
    writer_object = writer(f_object)
    for i in nrows:
        writer_object.writerow(i)
    f_object.close()

def splitnm():
    FirstName = []
    LastName = []
    fp = open(f'Updating/newdoc.csv')
    csvreader = csv.reader(fp)
    rows = []
    for row in csvreader:
        rows.append(row)
    # print(header)
    # print(rows)
    for line in rows: 
        # print(line[1])
        namestr = line[1].split(' ')
        if namestr[-1] == '':
            namestr = namestr[:-1]
        # print(namestr)
        FirstName.append(namestr[0])
        if namestr[-1] == 'Jr.' or namestr[-1] == 'I' or namestr[-1] == 'II' or namestr[-1] == 'III' or namestr[-1] == 'IV' or namestr[-1] == 'V' or namestr[-1] == 'Sr.':
            LastName.append(namestr[-2]+' '+namestr[-1])
        else:
            LastName.append(namestr[-1])
        # print(FirstName)
        # print(LastName)
    wb = load_workbook('Updating/namesplit.xlsx')
    ws = wb.active
    for i in range(1,335):
            ws.cell(row=i,column=1).value = FirstName[i-1]
            ws.cell(row=i,column=2).value = LastName[i-1]
    wb.save('Updating/namesplit.xlsx')

splitnm()