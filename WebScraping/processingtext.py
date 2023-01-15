import ast
from regex import P
import xlsxwriter
from openpyxl import load_workbook

def intoexcel():
    fp = open(f'results2.txt', encoding='utf8')
    wb = xlsxwriter.Workbook('try3.xlsx')
    ws = wb.add_worksheet()
    row = 0
    col = 0
    for line in fp:
        x = ast.literal_eval(line)
        for i in x:
            ws.write(col,row,i)
            row += 1
        col +=1
        row = 0
    wb.close()

intoexcel()


def splitnm():
    FirstName = []
    LastName = []
    fp = open(f'results2.txt', encoding='utf8')
    dcwb= xlsxwriter.Workbook('dc.xlsx')
    dcws = dcwb.add_worksheet('infosheet')
    col = 0
    row = 0
    for line in fp: 
        x = ast.literal_eval(line)
        namestr = x[0].split(' ')
        FirstName.append(namestr[0])
        if namestr[-1] == 'Jr.' or namestr[-1] == 'I' or namestr[-1] == 'II' or namestr[-1] == 'III' or namestr[-1] == 'IV' or namestr[-1] == 'V' or namestr[-1] == 'Sr.':
            LastName.append(namestr[-2]+' '+namestr[-1])
        else:
            LastName.append(namestr[-1])
        # print(FirstName)
        # print(LastName)
    for i in FirstName:
        dcws.write(col,row,i)
        col+=1
    col = 0
    for i in LastName:
        dcws.write(col,row+1,i)
        col+=1
    dcwb.close()

splitnm()

def stateandzipcode():
    states = []
    zipcodes = []
    fp = open(f'WebScraping/results2.txt', encoding='utf8')
    # existingWorksheet = dcwb.get_worksheet_by_name('infosheet')
    for line in fp:
        x = ast.literal_eval(line)
        try:
            stateandzipcodestr = x[1].split(' ')
            state = stateandzipcodestr[-2]
            zipcode = stateandzipcodestr[-1]
            # print(zipcode)
            states.append(state)
            zipcodes.append(zipcode)
        except:
            states.append('None')
            zipcodes.append('None')
    wb = load_workbook('dc.xlsx')
    ws = wb.active
    # for i in range(1,3134):
    #     ws.cell(row=i,column=3).value = states[i]
    for i in range(1,3138):
        # print(state)
        ws.cell(row=i,column=5).value = states[i-1]
        ws.cell(row=i,column=6).value = zipcodes[i-1]
    wb.save('dc.xlsx')

stateandzipcode()

def addressandcitynm():
    cities = []
    addresss = []
    fp = open(f'results2.txt', encoding='utf8')
    # existingWorksheet = dcwb.get_worksheet_by_name('infosheet')
    for line in fp:
        x = ast.literal_eval(line)
        try:
            infostr = x[1].split('^^')
            # print(infostr)
            citystr = infostr[-1].split(',')
            address = infostr[0] + infostr[1]
            print(address)
            print(citystr)
            citynm = citystr[0]
            # print(citynm)
            cities.append(citynm)
            addresss.append(address)
        except:
            cities.append('None')
            addresss.append('None')
    # print(len(cities))
    wb = load_workbook('dc.xlsx')
    ws = wb.active
    for i in range(1,3138):
        ws.cell(row=i,column=3).value = addresss[i-1]
        ws.cell(row=i,column=4).value = cities[i-1]
    wb.save('dc.xlsx')

addressandcitynm()