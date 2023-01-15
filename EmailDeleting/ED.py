import csv
from csv import writer
from openpyxl import load_workbook

        
#all contacts documents
def deleteables(docname):
    file = open(f'{docname}', encoding='utf8')
    csvreader = csv.reader(file)
    drows = [] #deleteable rows
    a = 0
    for row in csvreader:
        drows.append(row[0])
        a += 1
    print(f'total {a} in deleteables from {docname}')
    file.close()
    return drows


def alldoc():
    file = open(f'EmailDeleting/all.csv', encoding='utf8')
    csvreader = csv.reader(file)
    rows = [] #all rows

    num_of_lines_in_all = 0

    #count lines in all.csv
    for all in csvreader:
        rows.append(all)
        num_of_lines_in_all += 1
    print(f'total {num_of_lines_in_all} in the alldoc')
    file.close()
    return rows


def delete(rows, drows):
    c = 0
    d = 0
    #matching the deleteables
    for sublist in rows:
        # print(sublist)
        print(sublist[8])
        for email in drows:
            # print(email)
            if sublist[8] == email:
                try:
                    rows.remove(sublist)
                    d += 1
                except:
                    continue

    for i in rows:
        c +=1
    print(f'deleted {d} lines, {c} lines left in alldoc')
    return rows

drows = deleteables('EmailDeleting/all8.csv')
rows = alldoc()
new_row = delete(rows,drows) # do multiple times, reason is unknown

print(drows)
print(rows)
print(new_row) # at this time, row is already the deleted version bc of the remove function

# Open our existing CSV file in append mode
# Create a file object for this file
with open('new_list.csv', 'w') as f_object:
    writer_object = writer(f_object)
    for i in new_row:
        writer_object.writerow(i)
    f_object.close()

def splitnm():
    FirstName = []
    LastName = []
    fp = open(f'new_list.csv')
    csvreader = csv.reader(fp)
    header = []
    header = next(csvreader)
    rows = []
    for row in csvreader:
        rows.append(row)
    # print(header)
    # print(rows)
    for line in rows: 
        # print(line[1])
        namestr = line[1].split(' ')
        if namestr[-1] == '':
            namestr = namestr[:-1]
        # print(namestr)
        FirstName.append(namestr[0])
        if namestr[-1] == 'Jr.' or namestr[-1] == 'I' or namestr[-1] == 'II' or namestr[-1] == 'III' or namestr[-1] == 'IV' or namestr[-1] == 'V' or namestr[-1] == 'Sr.':
            LastName.append(namestr[-2]+' '+namestr[-1])
        else:
            LastName.append(namestr[-1])
        # print(FirstName)
        # print(LastName)
    wb = load_workbook('dc.xlsx')
    ws = wb.active
    for i in range(1,8435):
            ws.cell(row=i,column=1).value = FirstName[i-1]
    for i in range(1,8435):
            ws.cell(row=i,column=2).value = LastName[i-1]
    wb.save('dc.xlsx')

splitnm()
